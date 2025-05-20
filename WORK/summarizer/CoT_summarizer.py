from transformers import pipeline
from huggingface_model import HuggingFaceModel
import json 
import openpyxl
from openpyxl import load_workbook
import re
import torch
from summary_evaluator import LLMSummaryEvaluator
from NER import NER_Relationship
from transformers import BertTokenizer, BertModel
from bert_score import score


BERT_MODEL_DIR = "/home/dsta/saic/models/bert_model"
MODELS_DIR = "/home/dsta/pipeline-b/backend/models"
NER_TOKENIZER_PATH = "/home/dsta/saic/NER_models/ner_tokenizer"
EXCEL_FILE = "/home/dsta/saic/datasets/for_nsf_grading/B_100_articles_test_few_shot_and_dpo_entry.xlsx"
OUTPUT_EXCEL_FILE = "/home/dsta/saic/datasets/for_nsf_grading/eval_B_100_articles_test_few_shot_and_dpo_entry.xlsx"
COL_INDEX_FOR_CONTENT = 0 # INDEX means that counting is done 0,1,2,3...
COL_NUM_FOR_HUMAN_SUMMARY = 3 # NUM means that starting count is 1,2,3,4...
COL_INDEX_FOR_DATE = 3
COL_NUM_FOR_MACHINE_SUMMARY = 5
COL_NUM_FOR_ELEMENTS = 6
COL_NUM_FOR_NER_HUMAN_SUMM = 7
COL_NUM_FOR_NER_MACHINE_SUMM = 8
COL_NUM_FOR_BERT_SCORE = 9
COL_NUM_FOR_BERTSCORE_F1 = 10
MAX_NUM_COLS = 10
COY = 'B'
USE_DPO = True

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


class CustomSummarizer:
    def __init__(self):
        self.model = None
        self.model_type = None
        self.instruction_format_start = None
        self.instruction_format_end = None
        self.max_chunk_length = 250
        self.llm_summary_evaluator = None

    def unload_model(self):
        """Unload the current model."""
        self.model = None

    def load_summarization_model(self, model_name, tokenizer_name, checkpoint=None):
        """
        Load a custom summarization model from a checkpoint.
        This function sets an instruction format that helps control formatting.
        """
        self.unload_model()
        self.model = HuggingFaceModel(model_name, tokenizer_name)
        print("="*25, "model and tokenizer loaded! ", "="*25)
        if checkpoint:
            self.model.load_checkpoint(checkpoint)
        # Set an instruction format to guide the model for summarization.
        self.instruction_format_start = "<s>[INST] "
        self.instruction_format_end = "[/INST]"
        self.model.set_instructions(self.instruction_format_start, self.instruction_format_end)
        self.model_type = "summarization"
        print("Summarization model loaded")

    def chunk_text(self, text):
        """Split text into smaller chunks to avoid context overload."""
        sentences = re.split(r'(?<=[.!?])\s+', text.strip())
        chunks, current_chunk = [], []
        current_length = 0
        for sentence in sentences:
            sentence_length = len(sentence.split())
            if current_length + sentence_length > self.max_chunk_length:
                chunks.append(" ".join(current_chunk))
                current_chunk = [sentence]
                current_length = sentence_length
            else:
                current_chunk.append(sentence)
                current_length += sentence_length
        if current_chunk:
            chunks.append(" ".join(current_chunk))
        return chunks

    def ask_question(self, text, question, strict=True):
        """
        Ask the custom model a question about the text in a chain-of-thought manner.
        The prompt instructs the model to answer only using evidence from the document and output a JSON list.
        """
        print("="*20, f"In ask_question(): {question}", "="*20, "\n")
        # Create a prompt that asks a specific question about the document.
        prompt = (
            f"{self.instruction_format_start}"
            f"{question}\n"
            f"Document: {text}\n"
            "Answer using ONLY evidence from the document. "
            "Output a valid JSON list (e.g., [\"item1\", \"item2\"]). "
            "If no clear answer exists, return an empty list []. "
            "Do not invent or assume information."
            f"{self.instruction_format_end}"
        )
        answer = self.model.ask(prompt).strip()
        # Try to parse the JSON output
        try:
            parsed_answer = json.loads(answer)
            if isinstance(parsed_answer, list):
                return [str(item).strip() for item in parsed_answer if str(item).strip()]
            else:
                return []
        except json.JSONDecodeError:
            # Fallback: split the answer by commas if JSON parsing fails
            # return [item.strip() for item in answer.split(",") if item.strip()]
            return []


    def extract_elements(self, text):
        """
        Extract key elements with source verification.
        Returns a dictionary with source chunks as keys and extracted elements as values.
        """
        print("="*20, "In extract_element()", "="*20, "\n")
        
        # Dictionary to store elements by source chunk
        elements_by_source = {}
        
        # Process text in chunks
        chunks = self.chunk_text(text)
        
        # Define questions to extract information
        questions = {
            "entities": "List important names or entities mentioned in the document.",
            "events": "List key events explicitly described in the document.",
            "locations": "List locations explicitly mentioned in the document.",
            "results": "List outcomes or results of events explicitly stated in the document."
        }
        
        for chunk in chunks:
            # Initialize structure for this chunk if it doesn't exist yet
            if chunk not in elements_by_source:
                elements_by_source[chunk] = {
                    "entities": [],
                    "events": [],
                    "locations": [],
                    "results": []
                }
            
            # Process each question for this chunk
            for key, question in questions.items():
                # Ask question about the chunk
                answers = self.ask_question(chunk, question)
                
                # Add unique answers to the appropriate category
                for ans in answers:
                    if ans and ans not in elements_by_source[chunk][key]:
                        elements_by_source[chunk][key].append(ans)
        
        return elements_by_source

    def verify_element_sources(self, elements):
        """
        Print out the extracted elements with their source contexts.
        """
        print("=" * 20 + " Element Source Verification " + "=" * 20)
        
        for category, items in elements['_sources'].items():
            print(f"\n{category.upper()} Extraction Verification:")
            for item in items:
                print(f"- Element: {item['element']}")
                print(f"  Source Context: '...{item['source_chunk'][:300]}...'")
                print("-" * 50)

    

    def generate_summary(self, text, elements, date):
        """
        Generate a final summary using a chain-of-thought prompt that integrates the extracted elements.
        This prompt forces the model to only rely on the provided elements and the original text.
        """
        print("="*20, "In generate_summary()", "="*20, "\n")
        examples = [
            """On 24 Jun, it was reported that the MMEA activated a Search and Rescue (SAR) operation (Ops CARILAMAT) to locate a missing fisherman around 1.8nm west of Pantai Acheh Pulau Indah, Selangor on 23 Jun at 1700H but was postponed to 1930H due to bad weather and “dark conditions”. It was further reported that Ops CARILAMAT continued on 24 Jun at 0700H covering a search sector radius of 102nm2 which involved assets from “various agencies” such as: (i) the RMP Marine Police Force (MPF); (ii) Malaysian Fire and Rescue Department (JBPM); (iii) Malaysian Civil Defence Force (MCDF); and (iv) an RMP Air Operations Force (AOF) AgustaWestland AW139 helicopter. It was also reported that Selangor MMEA opened the CARILAMAT Forward Base at the Pulau Indah Marina Jetty to coordinate Ops CARILAMAT's efforts in the search sector. """,
            """On 2 Jul, Malaysian Army 3 Div Commander MG Zahari bin Mohd Ariffin chaired the Second Firepower Training 2024 (LKT 2024) Meeting at 3 Div HQ, Terendak Camp, Malacca. The meeting reportedly aimed to coordinate the conduct, administration, training and logistics plan for LKT 2024. Also present were Malaysian Army 1 Bde Commander BG Zamri bin Othman, Malaysian Army 4 Bde (Mech) Commander BG Wan Mohd Faizal Shahrin bin W. Mohamad Fablillah, Malaysian Army 7 Bde Commander BG Mohamad Suria bin Mohamad Saad and Malaysian Army 3 Div Artillery Commander BG Khairul Izan bin Muhammad Suffian, among others. """
        ]

        # dates = elements.get('dates', [])
        # primary_date = dates[0] if dates else "a specific date"
        entities = ", ".join(elements.get('entities', []) or ["unknown entities"])
        events = ", ".join(elements.get('events', []) or ["unspecified events"])
        locations = ", ".join(elements.get('locations', []) or ["unknown locations"])
        results = " ".join(elements.get('results', []) or ["No specific outcomes were reported."])

        # Create a chain-of-thought explanation that lays out the extracted details.
        chain_of_thought = (
            "Using ONLY the following extracted information:\n"
            f"Entities: {entities}\n"
            # f"Dates: {', '.join(dates)}\n"
            f"Date: {date}\n"
            f"Events: {events}\n"
            f"Locations: {locations}\n"
            f"Results: {results}\n\n"
            "Generate a concise summary STRICTLY following ONLY information from the above extracted elements."
            f"Your summary should always start with 'On {date}, it was reported that...' "
            f"""Summarize like these examples:
            Example 1: {examples[0]}
            Example 2: {examples[1]}"""
            "IGNORE: Opinions, historical background"
            "DO NOT USE INFORMATION FROM ANYWHERE ELSE EXCEPT THE EXTRACTED ELEMENTS."
            "Keep the summary to 2 to 5 sentences."
            f"Document: {text}"
        )
        # Wrap the chain-of-thought in the instruction format.
        prompt = f"{self.instruction_format_start}{chain_of_thought}{self.instruction_format_end}"
        summary = self.model.ask(prompt)

        return summary.strip()

    def evaluate_summary_consistency(self, original_text, summary):
        """
        Evaluate the factual consistency of the summary
        """
        # Use a factual consistency metric
        consistency_score = self._calculate_factual_consistency(
            original_text, 
            summary
        )
        
        return consistency_score

    def _calculate_factual_consistency(self, original_text, summary):
        """
        Calculate a factual consistency score
        """
        # Potential implementation using semantic similarity
        # or more advanced NLP techniques
        from sentence_transformers import SentenceTransformer
        import numpy as np
        
        model = SentenceTransformer('all-MiniLM-L6-v2')
        
        # Encode original text and summary
        original_embedding = model.encode(original_text)
        summary_embedding = model.encode(summary)
        
        # Calculate cosine similarity
        similarity = np.dot(original_embedding, summary_embedding) / (
            np.linalg.norm(original_embedding) * np.linalg.norm(summary_embedding)
        )
        
        return similarity

    def extract_source_evidences(self, elements):
        source_evidences = {}
        for category, items in elements['_sources'].items():
            print(f"\n{category.upper()} Extraction Verification:")
            for item in items:
                source_evidences[item['source_chunk']] = item['element'] 

    def generate_summary_reliability_report(self, elements):
        # Generate summary
        # summary = self.generate_summary(article_content, elements, date)
        
        # Extract source evidences
        source_evidences = self.extract_source_evidences(elements)
        
        # # Calculate factual consistency
        # consistency_score = self._calculate_factual_consistency(article_content, summary)
        
        # # Check coverage of key elements
        # element_coverage = self._check_element_coverage(elements, summary)
        
        return {
            # 'summary': summary,
            'source_evidences': source_evidences,
            # 'consistency_score': consistency_score,
            # 'element_coverage': element_coverage,
            # 'reliability_rating': self._calculate_overall_reliability(
            #     consistency_score, 
            #     element_coverage,
            #     source_evidences
            # )
        }
        

def model_selector(coy):
    if coy == 'A':
        return '/home/dsta/saic/models/A_Coy/A_mistral_with_lora_fine_tuned'
    if coy == 'B':
        return '/home/dsta/saic/mistral_with_lora_fine_tuned'
    if coy == 'C':
        return '/home/dsta/saic/models/C_Coy/C_mistral_with_lora_fine_tuned'
    if coy == 'D':
        return '/home/dsta/saic/models/D_Coy/D_mistral_with_lora_fine_tuned'

def get_dpo_path(coy):
    if coy == 'A':
        return '/home/dsta/saic/models/dpo_summarizer_large/A/checkpoint-288'
    if coy == 'B':
        return '/home/dsta/saic/models/dpo_summarizer_large/B/checkpoint-414'
    if coy == 'C':
        return ''
    if coy == 'D':
        return '/home/dsta/saic/models/dpo_summarizer_large/D/dpo_summarizer_200/checkpoint-69'
    
def get_ner_model_path(coy):
    if coy == 'A':
        return '/home/dsta/saic/NER_models/ACOYNERMODEL'
    if coy == 'B':
        return '/home/dsta/saic/NER_models/BCOYNERMODEL'
    if coy == 'C':
        return ''
    if coy == 'D':
        return '/home/dsta/saic/NER_models/DCOYNERMODEL'


def test_sentence(ner_model, sentence):
    """Test a single sentence"""
    ner_model_path = get_ner_model_path(COY)
    try:
        print("\nInitializing model and tokenizer...")
        ner_model.load(ner_model_path, NER_TOKENIZER_PATH)

        print("="*100)
        print(f"\nProcessing sentence: {sentence}")
        print("="*100)

        results = ner_model.infer(sentence)
        results_string = ""
        print("\nEntities found:")
        if not results['entities']:
            print("No entities found")
        else:
            for entity in results['entities']:
                print(f"- {entity['tag']}: {entity['text']}")
                results_string += f"{entity['tag']}: {entity['text']}\n"
            
        return results_string   
        
    except Exception as e:
        print(f"\nError processing sentence: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def get_bert_score(bert_model, machine_summary, human_summary):
    # bert_scorer = BERTScorer(
    #     model_type=bert_model,
    #     lang="en",
    #     rescale_with_baseline=True,
    #     use_fast_tokenizer=True,
    #     num_layers=None
    # )
    human_summary_list = [human_summary] 
    machine_summary_list = [machine_summary]
    P, R, F1 = score(human_summary_list, machine_summary_list, model_type=bert_model, lang="en", num_layers=10)
    bertscore_p = P.item()
    bertscore_r = R.item()
    bertscore_f1 = F1.item()
    print(f"BERTScore - P: {bertscore_p:.4f}, R: {bertscore_r:.4f}, F1: {bertscore_f1:.4f}")
    return bertscore_f1


def process_excel_file(summarizer, input_file, output_file, ner_model, bert_model):
    """
    Load the Excel file, process each article in the first column,
    generate a summary, and write the summary to the fourth column.
    """
    wb = load_workbook(input_file)
    ws = wb.active
    

    max_columns = ws.max_column
    if max_columns < MAX_NUM_COLS:
        print(f"Warning: The input file has only {max_columns} columns. Adding columns to make {MAX_NUM_COLS}.")
        for col_num in range(max_columns + 1, MAX_NUM_COLS + 1):
            # Add column headers if needed
            ws.cell(row=1, column=col_num).value = f"Column {col_num}"
    
    print("Max columns have been checked!")

    ws.cell(row=1, column=COL_NUM_FOR_NER_HUMAN_SUMM).value = "NER Results Human Summary"
    ws.cell(row=1, column=COL_NUM_FOR_NER_MACHINE_SUMM).value = "NER Results Machine Summary"
    ws.cell(row=1, column=COL_NUM_FOR_BERTSCORE_F1).value = "BERTScore F1"
    # No header row, min_row=1, else min_row=2
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Get the article text from the first column (index 0)
        article = row[COL_INDEX_FOR_CONTENT].value # To check which col the coy puts their article content
        if article and isinstance(article, str) and article.strip():
            print(f"Processing article {row_idx}/{ws.max_row}: {article[:60]}...")
            elements = summarizer.extract_elements(article)
            date = row[COL_INDEX_FOR_DATE].value
            print(f"DATE:::::::::::::::::::::::::: {date}")
            summary = summarizer.generate_summary(article, elements, date)
            # summary = row[COL_NUM_FOR_MACHINE_SUMMARY - 1].value ### TO USE ONLY WHEN MACHINE SUMMARY HAS ALREADY BEEN GENERATED
            print("="*30)
            print("SUMMARY:\n", summary)
            print("="*30)
            print("generating machine written NER results...")
            machine_ner_results = test_sentence(ner_model, summary)
            print("generating human written NER results...")
            human_ner_results = test_sentence(ner_model, row[COL_NUM_FOR_HUMAN_SUMMARY - 1].value)
            bert_score_f1 = get_bert_score(bert_model, summary, row[COL_NUM_FOR_HUMAN_SUMMARY - 1].value)
            # consistency_score = summarizer.evaluate_summary_consistency(article, summary)
            # ws.cell(row=row_idx, column=COL_NUM_FOR_ELEMENTS).value = str(consistency_score)
            
            # Write the summary to the 5th column (index 4)
            ws.cell(row=row_idx, column=COL_NUM_FOR_MACHINE_SUMMARY).value = str(summary)
            ws.cell(row=row_idx, column=COL_NUM_FOR_NER_MACHINE_SUMM).value = str(machine_ner_results)
            ws.cell(row=row_idx, column=COL_NUM_FOR_NER_HUMAN_SUMM).value = str(human_ner_results)
            ws.cell(row=row_idx, column=COL_NUM_FOR_BERTSCORE_F1).value = bert_score_f1
        else:
            print("Skipping empty or invalid article row.")

    wb.save(output_file)
    print(f"Summaries have been written to {output_file}")




# Example usage
if __name__ == "__main__":
    # Initialize custom summarizer
    summarizer = CustomSummarizer()

    # Get path for coy model
    path = model_selector(COY)
    DPO_PATH = get_dpo_path(COY)

    # Initialize NER model
    ner_model = NER_Relationship()

    # Initialize summary evaluator
    llm_summary_evaluator = LLMSummaryEvaluator(model_path=DPO_PATH, tokenizer_path=DPO_PATH)

    if not USE_DPO:
        print("**************** Not Using DPO. Using normal Mistral model with checkpoint. ***************")

        # Load your custom model with your checkpoint (update model names and checkpoint path as needed)
        summarizer.load_summarization_model(
            model_name=MODELS_DIR + "/sumgen/mistral_7b_model", 
            tokenizer_name=MODELS_DIR + "/sumgen/mistral_7b_tokenizer", 
            # checkpoint="/home/dsta/saic/mistral_with_lora_fine_tuned",
            # model_name=DPO_PATH,
            # tokenizer_name=DPO_PATH,
            checkpoint=path
            
        )
    else:
        print("**************** Using DPO ***************")
        summarizer.load_summarization_model(
            model_name=DPO_PATH,
            tokenizer_name=DPO_PATH,
        )
    process_excel_file(summarizer, EXCEL_FILE, OUTPUT_EXCEL_FILE, ner_model, BERT_MODEL_DIR)
